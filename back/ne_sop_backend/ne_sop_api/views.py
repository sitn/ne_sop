from ne_sop_api.models import (
    Document,
    Entity,
    EntityType,
    Event,
    EventType,
    Item,
    ItemStatus,
    ItemType,
    Template,
)
from ne_sop_api.serializers import (
    DocumentSerializer,
    EntitySerializer,
    EntityListSerializer,
    EntityTypeSerializer,
    NewItemSerializer,
    NestedItemSerializer,
    ItemListSerializer,
    ItemTypeSerializer,
    ItemStatusSerializer,
    EventSerializer,
    EventListSerializer,
    EventTypeSerializer,
    TemplateSerializer,
    UserSerializer,
    CurrentUserSerializer,
    ItemTypeStatisticsSerializer,
)

# from django.core.exceptions import PermissionDenied
from django.contrib.auth.models import User
from django.conf import settings
from django.db.models.functions import Lower, TruncYear
from django.db.models import Q, Count
from django.shortcuts import get_object_or_404
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from drf_spectacular.utils import extend_schema
from rest_framework.decorators import api_view, action
from rest_framework.parsers import MultiPartParser
from rest_framework import filters
from django.core.paginator import Paginator
from django.http import HttpResponse, HttpResponseForbidden, FileResponse

from ne_sop_api.utils import Utils
from ne_sop_api.permissions import (
    IsSuperuserPermission,
    IsManagerPermission,
    IsManagerOrReadOnlyPermission,
    # isAllowedRegadingEntitiesAndItemId,
)

from pathlib import PurePath
import os
import datetime
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from io import BytesIO


# %% TEST BACKEND
@api_view(["GET"])
def test_api(request):
    """
    Test backend
    """
    html = "<h1>Congratulations !</h1>"
    html += f"<h2>Hello {request.user}, It works</h2>"
    return HttpResponse(content=html, status=200)


# %% CURRENT USER
class CurrentUserViewSet(viewsets.ViewSet):
    """
    Curent Users viewset
    """

    queryset = User.objects.all()
    serializer_class = CurrentUserSerializer
    permission_classes = [IsManagerOrReadOnlyPermission]

    @extend_schema(
        responses=CurrentUserSerializer,
        tags=["Users"],
    )
    def list(self, request):
        queryset = User.objects.all()
        current_user = get_object_or_404(queryset, pk=request.user.id)
        serializer = CurrentUserSerializer(current_user)
        return Response(serializer.data)


# %% USER
class UserViewSet(viewsets.ViewSet):
    """
    Users viewset
    """

    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsManagerPermission]

    @extend_schema(
        responses=UserSerializer,
        tags=["Users"],
    )
    def list(self, request):
        serializer = UserSerializer(self.queryset, many=True)
        return Response(serializer.data)


# %% ENTITY TYPE
class EntityTypeViewSet(viewsets.ViewSet):
    """
    Entity types viewset
    """

    # queryset = EntityType.objects.all()
    # filter to keep "Parlementaire", "Groupe parlementaire", "Autre"
    # queryset = EntityType.objects.filter(id__in=[2, 3, 4])
    queryset = EntityType.objects.filter(service=False)
    serializer_class = EntityTypeSerializer

    @extend_schema(
        responses=EntityTypeSerializer,
        tags=["Entity type"],
    )
    def list(self, request):
        serializer = EntityTypeSerializer(self.queryset, many=True)
        return Response(serializer.data)


# %% PARLIAMENTARY TYPE
"""
class ParliamentaryTypeViewSet(viewsets.ViewSet):

    # queryset = EntityType.objects.all()
    queryset = EntityType.objects.filter(id__in=[2, 3, 4])
    serializer_class = EntityTypeSerializer

    @extend_schema(
        responses=EntityTypeSerializer,
        tags=["Entity type"],
    )
    def list(self, request):
        serializer = EntityTypeSerializer(self.queryset, many=True)
        return Response(serializer.data)
"""


# %% ENTITY
class EntityViewSet(viewsets.ViewSet):
    """
    Entities viewset
    """

    serializer_class = EntitySerializer
    search_fields = ["name", "email", "telephone"]
    permission_classes = [IsManagerOrReadOnlyPermission]

    @extend_schema(
        tags=["Entities"],
    )
    def list(self, request):
        filter = filters.SearchFilter()

        queryset = filter.filter_queryset(request, Entity.objects.all(), self)
        # queryset = filter.filter_queryset(request, Entity.objects.filter(type__in=[3]), self)

        # if self.request.user.is_superuser:
        #    queryset = filter.filter_queryset(request, Entity.objects.all(), self)
        # else:
        #    queryset = filter.filter_queryset(request, Entity.objects.filter(type__in=[2, 3, 4]), self)

        # queryset = filter.filter_queryset(request, Entity.objects.all(), self)
        # queryset = filter.filter_queryset(
        #     request, Entity.objects.filter(type__in=[2, 3, 4]), self
        # )

        # queryset = Entity.objects.all()
        name = request.query_params.get("name", "")
        type = request.query_params.get("type")
        service = request.query_params.get("service")
        page = int(request.query_params.get("page", "1"))
        size = int(request.query_params.get("size", "10"))
        sortby = request.query_params.get("sortby", "id")
        descending = request.query_params.get("descending", "false")
        # all_fields = Entity._meta.fields

        if sortby not in ["id", "name", "type"]:
            sortby = "id"

        if descending not in ["true", "false"]:
            descending = "false"

        if name is not None:
            queryset = queryset.filter(name__icontains=name)

        if type:
            queryset = queryset.filter(type__in=type.split(","))

        if service == "true":
            queryset = queryset.filter(type__service=True)

        if service == "false":
            queryset = queryset.filter(type__service=False)

        if descending == "true":
            paginator = Paginator(queryset.order_by(Lower(sortby).desc()), size)
        else:
            paginator = Paginator(queryset.order_by(Lower(sortby).asc()), size)

        queryset = paginator.page(page)
        nrows = paginator.count
        npages = paginator.num_pages

        serializer = EntityListSerializer(queryset, many=True)

        return Response(
            {
                "page": page,
                "npages": npages,
                "nrows": nrows,
                "sortby": sortby,
                "descending": descending,
                "results": serializer.data,
            }
        )

    @extend_schema(
        tags=["Entities"],
    )
    def create(self, request):
        serializer = EntitySerializer(data=request.data)
        if serializer.is_valid():
            # print("CREATE ENTITY")
            # print(serializer.validated_data.get("type").service)

            if (not self.request.user.is_superuser) & serializer.validated_data.get("type").service:
                # if not self.request.user.is_superuser & serializer.validated_data.get("type").id not in [2, 3, 4]:
                return Response(
                    {"msg": "You do not have permissions to create a service"},
                    status=status.HTTP_403_FORBIDDEN,
                )
            else:
                serializer.save()

            # return Response({"msg": "New entity created"}, status=status.HTTP_201_CREATED)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        tags=["Entities"],
    )
    def retrieve(self, request, pk=None):
        queryset = Entity.objects.all()
        entity = get_object_or_404(queryset, pk=pk)
        serializer = EntitySerializer(entity)
        return Response(serializer.data)

    # @extend_schema(
    #     tags=["Entities"],
    # )
    # def update(self, request, pk=None):
    #     queryset = Entity.objects.all()
    #     entity = get_object_or_404(queryset, pk=pk)
    #     serializer = EntitySerializer(entity, data=request.data)
    #     # print("UPDATE ENTITY")
    #     # print(serializer.validated_data.get("type").service)

    #     if serializer.is_valid():

    #         # if (not self.request.user.is_superuser) & serializer.validated_data.get("type").service:
    #         if (not self.request.user.is_superuser) & entity.type.service:
    #             # if not self.request.user.is_superuser & serializer.validated_data.get("type").id not in [2, 3, 4]:
    #             return Response(
    #                 {"msg": "You do not have permissions to update a service"},
    #                 status=status.HTTP_403_FORBIDDEN,
    #             )
    #         else:
    #             serializer.save()
    #             return Response(serializer.data)
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        tags=["Entities"],
    )
    def partial_update(self, request, *args, **kwargs):

        # instance = self.queryset.get(pk=kwargs.get('pk'))

        queryset = Entity.objects.all()
        # entity = get_object_or_404(queryset, pk=pk)
        entity = get_object_or_404(queryset, pk=kwargs.get("pk"))
        serializer = EntitySerializer(entity, data=request.data, partial=True)
        # print("UPDATE ENTITY")
        # print(serializer.validated_data.get("type").service)

        if serializer.is_valid():

            # if (not self.request.user.is_superuser) & serializer.validated_data.get("type").service:
            if (not self.request.user.is_superuser) & entity.type.service:
                # if not self.request.user.is_superuser & serializer.validated_data.get("type").id not in [2, 3, 4]:
                return Response(
                    {"msg": "You do not have permissions to update a service"},
                    status=status.HTTP_403_FORBIDDEN,
                )
            else:
                serializer.save()
                return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        tags=["Entities"],
    )
    def destroy(self, request, pk=None):
        queryset = Entity.objects.all()
        entity = get_object_or_404(queryset, pk=pk)

        # print("DESTROY ENTITY")
        # print(entity.type.name)
        # print(entity.type.service)
        # print(self.request.user)
        # print(self.request.user.is_superuser)

        if (not self.request.user.is_superuser) & entity.type.service:
            # if not self.request.user.is_superuser & entity.type not in [2, 3, 4]:
            return Response(
                {"msg": "You do not have permission to delete a service"},
                status=status.HTTP_403_FORBIDDEN,
            )
        else:
            entity.delete()
            # print("DELETED")

        return Response({"msg": "Entity deleted"})


# %% SERVICE
"""
class ServiceViewSet(viewsets.ViewSet):

    serializer_class = EntitySerializer
    search_fields = ["name", "email", "telephone"]
    permission_classes = [IsSuperuserPermission]

    @extend_schema(
        tags=["Entities"],
    )
    def list(self, request):
        filter = filters.SearchFilter()
        queryset = filter.filter_queryset(request, Entity.objects.filter(type=1), self)

        # queryset = Entity.objects.all()
        name = request.query_params.get("name", "")
        page = int(request.query_params.get("page", "1"))
        size = int(request.query_params.get("size", "10"))
        sortby = request.query_params.get("sortby", "id")
        descending = request.query_params.get("descending", "false")
        # all_fields = Entity._meta.fields

        if sortby not in ["id", "name"]:
            sortby = "id"

        if descending not in ["true", "false"]:
            descending = "false"

        if name is not None:
            queryset = queryset.filter(name__icontains=name)

        if descending == "true":
            paginator = Paginator(queryset.order_by(Lower(sortby).desc()), size)
        else:
            paginator = Paginator(queryset.order_by(Lower(sortby).asc()), size)

        queryset = paginator.page(page)
        nrows = paginator.count
        npages = paginator.num_pages

        serializer = EntityListSerializer(queryset, many=True)

        return Response(
            {
                "page": page,
                "npages": npages,
                "nrows": nrows,
                "sortby": sortby,
                "descending": descending,
                "results": serializer.data,
            }
        )

    @extend_schema(
        tags=["Entities"],
    )
    def create(self, request):
        serializer = EntitySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            # return Response({"msg": "New entity created"}, status=status.HTTP_201_CREATED)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        tags=["Entities"],
    )
    def retrieve(self, request, pk=None):
        queryset = Entity.objects.all()
        entity = get_object_or_404(queryset, pk=pk)
        serializer = EntitySerializer(entity)
        return Response(serializer.data)

    @extend_schema(
        tags=["Entities"],
    )
    def update(self, request, pk=None):
        queryset = Entity.objects.all()
        entity = get_object_or_404(queryset, pk=pk)
        serializer = EntitySerializer(entity, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        tags=["Entities"],
    )
    def destroy(self, request, pk=None):
        queryset = Entity.objects.all()
        entity = get_object_or_404(queryset, pk=pk)
        entity.delete()
        return Response({"msg": "Service deleted"})
"""


# %% ITEM TYPE
class ItemTypeViewSet(viewsets.ViewSet):
    """
    Item types viewset
    """

    queryset = ItemType.objects.all()
    serializer_class = ItemTypeSerializer

    @extend_schema(
        responses=ItemTypeSerializer,
        tags=["Item types"],
    )
    def list(self, request):
        serializer = ItemTypeSerializer(self.queryset, many=True)
        return Response(serializer.data)


# %% ITEM STATUS
class ItemStatusViewSet(viewsets.ViewSet):
    """
    Item status viewset
    order_by(Lower(sortby).asc())
    """

    # order queryset by ascending id
    sortby = "id"
    queryset = ItemStatus.objects.all().order_by(Lower(sortby).asc())
    serializer_class = ItemStatusSerializer

    @extend_schema(
        responses=ItemStatusSerializer,
        tags=["Item status"],
    )
    def list(self, request):
        serializer = ItemStatusSerializer(self.queryset, many=True)
        return Response(serializer.data)


# ITEM SUMMARY
class ItemSummaryViewSet(viewsets.ViewSet):
    """
    Item summary viewset
    """

    queryset = Item.objects.all()
    serializer_class = NestedItemSerializer
    search_fields = ["title", "number"]

    @extend_schema(
        tags=["Items"],
    )
    def retrieve(self, request, pk=None):
        item = get_object_or_404(self.queryset, pk=pk)
        serializer = NestedItemSerializer(item)
        return Response(serializer.data)


# ITEM
class ItemViewSet(viewsets.ViewSet):
    """
    Item viewset
    """

    serializer_class = NewItemSerializer
    search_fields = ["title", "number"]

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(name="Manager").exists():
            return Item.objects.all()
        return Item.objects.filter(Q(lead__in=user.entities.all()) | Q(support__in=user.entities.all())).distinct()
        # return Item.objects.all()

    @extend_schema(
        responses=NewItemSerializer,
        tags=["Items"],
    )
    def list(self, request):
        search_filter = filters.SearchFilter()
        queryset = search_filter.filter_queryset(request, self.get_queryset(), self)

        number = request.query_params.get("number")
        title = request.query_params.get("title")
        item_type = request.query_params.get("type")
        status = request.query_params.get("status")
        service = request.query_params.get("service")
        page = int(request.query_params.get("page", "1"))
        size = int(request.query_params.get("size", "10"))
        sortby = request.query_params.get("sortby", "id")
        descending = request.query_params.get("descending", "false")

        # all_fields = Entity._meta.fields
        if sortby not in [
            "id",
            "number",
            "title",
            "type",
            "status",
            "urgent",
            "startdate",
            "enddate",
        ]:
            sortby = "id"

        if descending not in ["true", "false"]:
            descending = "false"

        if number:
            queryset = queryset.filter(number=number)  # __icontains

        if title:
            queryset = queryset.filter(title=title)  # __icontains

        if status:
            queryset = queryset.filter(status__id__in=list(filter(None, status.split(",")))).distinct()

        if item_type:
            queryset = queryset.filter(type__id__in=list(filter(None, item_type.split(",")))).distinct()

        if service:
            # queryset = queryset.filter(lead__id__in=list(filter(None, service.split(","))))
            queryset = queryset.filter(Q(lead__id__in=list(filter(None, service.split(",")))) | Q(support__id__in=list(filter(None, service.split(","))))).distinct()

            # Item.objects.filter(Q(lead__in=user.entities.all()) | Q(support__in=user.entities.all()))

        if descending == "true":
            paginator = Paginator(queryset.order_by(Lower(sortby).desc()), size)
        else:
            paginator = Paginator(queryset.order_by(Lower(sortby).asc()), size)

        queryset = paginator.page(page)
        nrows = paginator.count
        npages = paginator.num_pages

        serializer = ItemListSerializer(queryset, many=True)

        # return Response(serializer.data)
        return Response(
            {
                "page": page,
                "npages": npages,
                "nrows": nrows,
                "sortby": sortby,
                "descending": descending,
                "results": serializer.data,
            }
        )

    @extend_schema(
        tags=["Items"],
    )
    def create(self, request):
        serializer = NewItemSerializer(data=request.data)
        if serializer.is_valid():
            item = serializer.save()
            Utils.itemCreatedNotification(item, request)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        # print("serializer.errors")
        # print(serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        tags=["Items"],
    )
    def retrieve(self, request, pk=None):
        item = get_object_or_404(self.get_queryset(), pk=pk)
        serializer = NewItemSerializer(item)
        return Response(serializer.data)

    @extend_schema(
        tags=["Items"],
    )
    def update(self, request, pk=None):
        item = get_object_or_404(self.get_queryset(), pk=pk)
        serializer = NewItemSerializer(item, data=request.data)
        if serializer.is_valid():
            item = serializer.save()
            if item.autonotify is True:
                Utils.itemChangedNotification(item, request)
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        tags=["Items"],
    )
    def destroy(self, request, pk=None):
        item = get_object_or_404(self.get_queryset(), pk=pk)
        Utils.itemRemovedNotification(item, request)
        item.delete()
        return Response({"msg": "Item deleted"})

    # def list(self, request):
    @action(detail=False, methods=["get"], url_path="download")
    def export_excel(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        now = datetime.datetime.now()
        filename = f'{datetime.datetime.strftime(now, "%Y%m%d-%H%M%S")}_ObjetsParlementaires.xlsx'

        ## Save results in Excel file
        # create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Objets parlementaires"
        ws.cell(1, 1).value = "Liste des objets parlementaires"
        ws.cell(1, 1).font = Font(size=16, bold=True)
        ws.cell(2, 1).value = "Date"
        ws.cell(2, 2).value = datetime.datetime.strftime(now, "%d.%m.%Y")

        row_id = 4

        # table header
        ws.cell(row_id, 1).value = "Uuid"
        ws.cell(row_id, 2).value = "Numéro"
        ws.cell(row_id, 3).value = "Titre"
        ws.cell(row_id, 4).value = "Création"
        ws.cell(row_id, 5).value = "Auteur"
        ws.cell(row_id, 6).value = "Type"
        ws.cell(row_id, 7).value = "Statut"
        ws.cell(row_id, 8).value = "Description"
        ws.cell(row_id, 9).value = "Urgent"
        ws.cell(row_id, 10).value = "Réponse écrite"
        ws.cell(row_id, 11).value = "Réponse orale"
        ws.cell(row_id, 12).value = "Date de dépôt"
        ws.cell(row_id, 13).value = "Date de retour au SG"
        ws.cell(row_id, 14).value = "Notifications auto"
        ws.cell(row_id, 15).value = "Valide"
        ws.cell(row_id, 16).value = "Retard"
        ws.cell(row_id, 17).value = "Service principal"
        ws.cell(row_id, 18).value = "Service(s) en appui"
        ws.cell(row_id, 19).value = "Événements"
        ws.cell(row_id, 20).value = "Documents"

        sep = "\n"
        for op in queryset:
            row_id += 1
            # table row
            ws.cell(row_id, 1).value = str(op.uuid) or None
            ws.cell(row_id, 2).value = op.number or None
            ws.cell(row_id, 3).value = op.title or None
            ws.cell(row_id, 4).value = datetime.datetime.strftime(op.created, "%d.%m.%Y %H:%M:%S") if op.created is not None else None
            ws.cell(row_id, 5).value = op.author.name or None
            ws.cell(row_id, 6).value = op.type.name or None
            ws.cell(row_id, 7).value = op.status.name or None
            ws.cell(row_id, 8).value = op.description or None
            ws.cell(row_id, 9).value = op.urgent or None
            ws.cell(row_id, 10).value = op.writtenresponse or None
            ws.cell(row_id, 11).value = op.oralresponse or None
            ws.cell(row_id, 12).value = datetime.datetime.strftime(op.startdate, "%d.%m.%Y") if op.startdate is not None else None
            ws.cell(row_id, 13).value = datetime.datetime.strftime(op.enddate, "%d.%m.%Y") if op.enddate is not None else None
            ws.cell(row_id, 14).value = op.autonotify or None
            ws.cell(row_id, 15).value = op.valid or None
            ws.cell(row_id, 16).value = op.late or None
            ws.cell(row_id, 17).value = op.lead.name or None
            ws.cell(row_id, 18).value = sep.join([sup.name for sup in op.support.all()])
            ws.cell(row_id, 19).value = sep.join([tmp.type.name for tmp in op.events.all()])
            ws.cell(row_id, 20).value = sep.join([tmp.filename for tmp in op.documents.all()])
            ws.cell(row_id, 18).alignment = Alignment(wrap_text=True)
            ws.cell(row_id, 19).alignment = Alignment(wrap_text=True)
            ws.cell(row_id, 20).alignment = Alignment(wrap_text=True)

            # adjust row height
            ws.row_dimensions[row_id].height = (max([len(op.support.all()), len(op.events.all()), len(op.events.all())]) - 1) * 15 + 15

        # adjust column width
        ws = Utils.auto_adjust_excel_column_width(ws)

        # create table in worksheet
        table = Table(displayName="Table1", ref="A4:" + get_column_letter(20) + str(row_id))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        response = FileResponse(file_stream, filename=filename, as_attachment=True)
        return response


# %% EVENT TYPE
class EventTypeViewSet(viewsets.ViewSet):
    """
    Event types viewset
    """

    # order queryset by ascending id
    sortby = "id"
    queryset = EventType.objects.all().order_by(Lower(sortby).asc())
    serializer_class = EventTypeSerializer

    @extend_schema(
        responses=EventTypeSerializer,
        tags=["Event types"],
    )
    def list(self, request):
        serializer = EventTypeSerializer(self.queryset, many=True)
        return Response(serializer.data)


# %% EVENT
class EventViewSet(viewsets.ViewSet):
    """
    Events viewset
    """

    serializer_class = EventSerializer
    search_fields = ["date", "item__number", "item__title"]
    permission_classes = [IsManagerOrReadOnlyPermission]

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(name="Manager").exists():
            return Event.objects.all()

        return Event.objects.filter(Q(item__lead__in=user.entities.all()) | Q(item__support__in=user.entities.all()))

    @extend_schema(
        responses=EventSerializer,
        tags=["Events"],
    )
    def list(self, request):
        filter = filters.SearchFilter()
        queryset = filter.filter_queryset(request, self.get_queryset(), self)

        # queryset = Event.objects.all()
        item = request.query_params.get("item", None)
        page = int(request.query_params.get("page", "1"))
        size = int(request.query_params.get("size", "10"))
        sortby = request.query_params.get("sortby", "id")
        descending = request.query_params.get("descending", "false")

        if sortby not in ["id", "date", "type", "item"]:
            sortby = "date"

        if descending not in ["true", "false"]:
            descending = "true"

        if item and len(item) > 0:
            queryset = queryset.filter(item__exact=item)

        if descending == "true":
            paginator = Paginator(queryset.order_by(Lower(sortby).desc()), size)
        else:
            paginator = Paginator(queryset.order_by(Lower(sortby).asc()), size)

        queryset = paginator.page(page)
        nrows = paginator.count
        npages = paginator.num_pages

        serializer = EventListSerializer(queryset, many=True)

        return Response(
            {
                "page": page,
                "npages": npages,
                "nrows": nrows,
                "sortby": sortby,
                "descending": descending,
                "results": serializer.data,
            }
        )

    @extend_schema(
        tags=["Events"],
    )
    def create(self, request):
        serializer = EventSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"msg": "Event created"}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        tags=["Events"],
    )
    def retrieve(self, request, pk=None):
        event = get_object_or_404(self.get_queryset(), pk=pk)
        serializer = EventSerializer(event)
        return Response(serializer.data)

    @extend_schema(
        tags=["Events"],
    )
    def update(self, request, pk=None):
        event = get_object_or_404(self.get_queryset(), pk=pk)
        serializer = EventSerializer(event, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        tags=["Events"],
    )
    def destroy(self, request, pk=None):
        event = get_object_or_404(self.get_queryset(), pk=pk)
        event.delete()
        return Response({"msg": "Event deleted"})


# %% TEMPLATE
class TemplateViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Templates viewset
    """

    queryset = Template.objects.all()
    serializer_class = TemplateSerializer


class TemplateTypeViewSet(viewsets.ViewSet):
    """
    Template types viewset
    """

    queryset = Template.objects.all()
    serializer_class = TemplateSerializer

    @extend_schema(
        responses=TemplateSerializer,
        tags=["Template type"],
    )
    def list(self, request):
        itemtype_id = request.query_params.get("itemtype_id") if "itemtype_id" in request.query_params else None

        templates = Template.objects
        if itemtype_id is not None:
            templates = templates.filter(item_types__id=itemtype_id)
        templates = templates.filter(valid=True).all()

        serializer = TemplateSerializer(templates, many=True)

        return Response(serializer.data)


class DocumentViewSet(viewsets.ViewSet):
    """
    New document viewset
    """

    parser_classes = [MultiPartParser]
    serializer_class = DocumentSerializer

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(name="Manager").exists():
            return Document.objects.all()
        return Document.objects.filter(Q(item__lead__in=user.entities.all()) | Q(item__support__in=user.entities.all()))

    @extend_schema(
        responses=DocumentSerializer,
        tags=["Document"],
    )
    def create(self, request):
        serializer = DocumentSerializer(data=request.data, context={"request": self.request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        responses=DocumentSerializer,
        tags=["Document"],
    )
    def retrieve(self, request, pk):
        document = get_object_or_404(self.get_queryset(), pk=pk)

        filepath = PurePath(settings.MEDIA_ROOT, document.file.name)
        filepath = open(filepath, "rb")

        response = FileResponse(filepath, filename=document.filename, as_attachment=True)

        headers = response.headers
        headers["Content-Type"] = "application/download"
        headers["Accept-Ranges"] = "bite"
        response["Content-Disposition"] = f"attachment; filename={document.filename}"
        return response

    @extend_schema(
        responses=DocumentSerializer,
        tags=["Document"],
    )
    def destroy(self, request, pk=None):
        document = get_object_or_404(self.get_queryset(), pk=pk)
        os.remove(PurePath(settings.MEDIA_ROOT, document.file.name))
        document.delete()
        return Response({"msg": "Document deleted"})


# %% UPDATE LATE ATTRIBUTE IN ITEMS
class LateItemsViewSet(viewsets.ViewSet):
    """
    Late item viewset
    """

    # serializer_class = CurrentUserSerializer
    permission_classes = [IsManagerOrReadOnlyPermission]

    @extend_schema(
        responses=CurrentUserSerializer,
        tags=["Items"],
    )
    def list(self, request):

        key = request.query_params.get("key", None)

        if key != settings.SECRET_UUID:
            return HttpResponseForbidden()

        # filter items that are late today
        # queryset = Item.objects.filter(enddate=datetime.date.today() - datetime.timedelta(days=1), status__in=[1, 2])
        queryset = Item.objects.filter(enddate=datetime.date.today() - datetime.timedelta(days=1), status__deadline=True)
        queryset.update(late=True)

        for item in queryset:
            # print(serializer.data)
            Utils.itemLateNotification(item, request)

        return Response(
            {
                "msg": f"{queryset.count()} late item(s) updated",
            }
        )


# %% ITEM STATISTICS
class ItemTypeStatisticsDepositedViewSet(viewsets.ViewSet):
    """
    Item deposited statistics viewset
    """

    queryset = ItemType.objects.all()
    serializer_class = ItemTypeStatisticsSerializer
    # permission_classes = [IsManagerPermission]

    @extend_schema(
        responses=ItemTypeStatisticsSerializer,
        tags=["Statistics"],
    )
    def list(self, request):
        queryset = ItemType.objects.annotate(year=TruncYear("item__startdate")).annotate(num_items=Count("item"))

        # get unique set of years
        years = list(set(int(x.year.year) if x.year is not None else None for x in queryset))
        years.remove(None)
        years.sort(reverse=True)

        # get unique set of itemtypes
        itemtypes = list(set(x.name for x in queryset))
        itemtypes.sort()

        # prepare result object
        result = []
        for year in years:
            tmp = {"year": year}
            for itemtype in itemtypes:
                for qs in queryset:
                    if qs.year is not None and qs.year.year == year and qs.name == itemtype:
                        tmp[itemtype] = qs.num_items
                        break
                    else:
                        tmp[itemtype] = 0
            result.append(tmp)

        return Response(result)


class ItemTypeStatisticsTreatedViewSet(viewsets.ViewSet):
    """
    Item treated statistics viewset
    """

    queryset = ItemType.objects.all()
    serializer_class = ItemTypeStatisticsSerializer
    # permission_classes = [IsManagerPermission]

    @extend_schema(
        responses=ItemTypeStatisticsSerializer,
        tags=["Statistics"],
    )
    def list(self, request):
        queryset = ItemType.objects.annotate(year=TruncYear("item__enddate")).annotate(num_items=Count("item"))

        # get unique set of years
        years = list(set(int(x.year.year) if x.year is not None else None for x in queryset))
        years.remove(None)
        years.sort(reverse=True)

        # get unique set of itemtypes
        itemtypes = list(set(x.name for x in queryset))
        itemtypes.sort()

        # prepare result object
        result = []
        for year in years:
            tmp = {"year": year}
            for itemtype in itemtypes:
                for qs in queryset:
                    if qs.year is not None and qs.year.year == year and qs.name == itemtype:
                        tmp[itemtype] = qs.num_items
                        break
                    else:
                        tmp[itemtype] = 0
            result.append(tmp)

        return Response(result)


class ServiceStatisticsViewSet(viewsets.ViewSet):
    """
    Item treated by service viewset
    """

    queryset = Entity.objects.all()
    # serializer_class = ItemTypeStatisticsSerializer
    # permission_classes = [IsManagerPermission]

    @extend_schema(
        # responses=ItemTypeStatisticsSerializer,
        tags=["Statistics"],
    )
    def list(self, request):
        queryset = Entity.objects.filter(type__service=True).prefetch_related("item__lead").annotate(year=TruncYear("item__startdate"))

        # get unique set of years
        years = list(set(int(x.year.year) if x.year is not None else None for x in queryset))
        years.remove(None)
        years.sort(reverse=True)

        # get unique set of services
        services = list(set(x.abbreviation for x in queryset))
        services.sort()

        # prepare result object
        result = []
        for year in years:
            tmp = {"year": year}
            for service in services:

                tmp[service] = len(Item.objects.filter(lead__abbreviation=service, startdate__year=year))

            result.append(tmp)

        return Response(result)


class StatutStatisticsViewSet(viewsets.ViewSet):
    """
    Item status statistics viewset
    """

    queryset = ItemStatus.objects.all()
    # serializer_class = ItemTypeStatisticsSerializer
    # permission_classes = [IsManagerPermission]

    @extend_schema(
        # responses=ItemTypeStatisticsSerializer,
        tags=["Statistics"],
    )
    def list(self, request):
        queryset = ItemStatus.objects.annotate(year=TruncYear("item__startdate")).annotate(num_items=Count("item"))

        # get unique set of years
        years = list(set(int(x.year.year) if x.year is not None else None for x in queryset))
        years.remove(None)
        years.sort(reverse=True)

        # get unique set of statuts
        statuts = list(set(x.name for x in queryset))
        statuts.sort()

        # prepare result object
        result = []
        for year in years:
            tmp = {"year": year}
            for statut in statuts:
                for qs in queryset:
                    if qs.year is not None and qs.year.year == year and qs.name == statut:
                        tmp[statut] = qs.num_items
                        break
                    else:
                        tmp[statut] = 0
            result.append(tmp)

        return Response(result)


class UrgentWrittenStatisticsViewSet(viewsets.ViewSet):
    """
    Item urgency statistics viewset
    """

    queryset = ItemStatus.objects.all()
    # serializer_class = ItemTypeStatisticsSerializer
    # permission_classes = [IsManagerPermission]

    @extend_schema(
        # responses=ItemTypeStatisticsSerializer,
        tags=["Statistics"],
    )
    def list(self, request):
        queryset = Item.objects.annotate(year=TruncYear("startdate"))
        # queryset_urgent = Item.objects.filter(urgent=True).annotate(year=TruncYear("startdate"))
        # queryset_writtenresponse = Item.objects.filter(writtenresponse=True).annotate(year=TruncYear("startdate"))

        # get unique set of years
        years = list(set(int(x.year.year) if x.year is not None else None for x in queryset))
        if None in years:
            years.remove(None)
        years.sort(reverse=True)

        # get unique set of statuts
        statuts = ["Urgence demandée", "Réponse écrite demandée"]

        # prepare result object
        result = []
        for year in years:
            tmp = {"year": year}
            for statut in statuts:
                if statut == "Urgence demandée":

                    tmp[statut] = len(Item.objects.filter(urgent=True, startdate__year=year))

                if statut == "Réponse écrite demandée":

                    tmp[statut] = len(Item.objects.filter(writtenresponse=True, startdate__year=year))

            result.append(tmp)

        return Response(result)
